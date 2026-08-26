from __future__ import annotations

import unittest
from decimal import Decimal
from pathlib import Path
import tempfile

import pdfplumber
from openpyxl import load_workbook

from scripts.ata_sessao_reports.data_normalizer import normalize_report_data
from scripts.ata_sessao_reports.excel import write_reports_workbooks
from scripts.ata_sessao_reports.models import AtaSessaoParseResult, LotItemData, LotRecord
from scripts.ata_sessao_reports.pdf_renderer import (
    write_ata_institucional_pdf,
    write_report_pdfs,
)
from scripts.ata_sessao_reports.reconciliation import reconcile_estimated_values
from scripts.ata_sessao_reports.sd_parser import SDItem, SDMetadata, SDRecord


def _sd_item(
    numero: int,
    descricao: str,
    *,
    catmat: str | None = None,
    quantidade: str = "1",
    unidade: str = "UN",
    unitario: str = "10.00",
) -> SDItem:
    quantity = Decimal(quantidade)
    unit_value = Decimal(unitario)
    return SDItem(
        numero=numero,
        catmat_catser=catmat,
        descricao=descricao,
        quantidade=quantity,
        percentual=Decimal("100"),
        unidade=unidade,
        preco_unitario=unit_value,
        preco_total=quantity * unit_value,
    )


def _sd(*items: SDItem) -> SDRecord:
    return SDRecord(
        source_path="sd.pdf",
        metadata=SDMetadata(
            numero_sd="152/2026",
            data_emissao=None,
            centro_custo=None,
            unidade_orcamentaria=None,
            elemento_despesa=None,
            fonte_recurso=None,
            valor_total=sum((item.preco_total for item in items), Decimal("0")),
            assunto_objeto=None,
            processo_administrativo="PA-152/2026",
        ),
        itens=list(items),
    )


def _item(
    description: str,
    *,
    number: str = "1",
    catmat: str | None = None,
    quantity: float = 1,
    unit: str = "UND",
) -> LotItemData:
    return LotItemData(
        item_numero=number,
        catmat_catser=catmat,
        descricao=description,
        quantidade=quantity,
        unidade=unit,
    )


class SDReconciliationTests(unittest.TestCase):
    def test_local_item_one_is_not_reused_across_failed_lots(self) -> None:
        result = AtaSessaoParseResult(
            source_path="ata.pdf",
            generated_at="2026-08-12T10:00:00",
            lotes=[
                LotRecord(1, "FRACASSADO", "Cadeira", itens=[_item("CADEIRA ESCOLAR")]),
                LotRecord(2, "DESERTO", "Mesa", itens=[_item("MESA ESCOLAR")]),
            ],
        )
        metadata = reconcile_estimated_values(
            result,
            _sd(
                _sd_item(1, "CADEIRA ESCOLAR", unitario="100.00"),
                _sd_item(2, "MESA ESCOLAR", unitario="250.00"),
            ),
        )

        self.assertEqual(result.lotes[0].itens[0].valor_unitario_estimado, 100.0)
        self.assertEqual(result.lotes[1].itens[0].valor_unitario_estimado, 250.0)
        self.assertEqual(result.lotes[0].itens[0].valor_estimado_correspondencia, "LOTE")
        self.assertEqual(result.lotes[1].itens[0].valor_estimado_correspondencia, "LOTE")
        self.assertEqual(metadata["fully_matched_lots"], 2)
        self.assertEqual(metadata["matched_items"], 2)
        self.assertEqual(metadata["unmatched_lots"], [])

    def test_unique_catmat_matches_and_duplicate_catmat_is_ambiguous(self) -> None:
        unique_result = AtaSessaoParseResult(
            source_path="ata.pdf",
            generated_at="2026-08-12T10:00:00",
            lotes=[
                LotRecord(
                    99,
                    "CANCELADO",
                    "Tubo",
                    itens=[_item("TUBO PVC 100 MM", catmat="000123456")],
                )
            ],
        )
        unique_metadata = reconcile_estimated_values(
            unique_result,
            _sd(_sd_item(7, "TUBO PVC 100 MM", catmat="123456", unitario="42.50")),
        )
        self.assertEqual(unique_metadata["matched_items"], 1)
        self.assertEqual(
            unique_result.lotes[0].itens[0].valor_estimado_correspondencia,
            "CATMAT_CATSER",
        )

        duplicate_result = AtaSessaoParseResult(
            source_path="ata.pdf",
            generated_at="2026-08-12T10:00:00",
            lotes=[
                LotRecord(
                    99,
                    "CANCELADO",
                    "Tubo",
                    itens=[_item("TUBO PVC 100 MM", catmat="123456")],
                )
            ],
        )
        duplicate_metadata = reconcile_estimated_values(
            duplicate_result,
            _sd(
                _sd_item(7, "TUBO PVC 100 MM", catmat="123456"),
                _sd_item(8, "TUBO PVC 100 MM REFORCADO", catmat="123456"),
            ),
        )
        self.assertEqual(duplicate_metadata["ambiguous_items"], 1)
        self.assertEqual(duplicate_metadata["ambiguous_lots"], [99])
        self.assertIsNone(duplicate_result.lotes[0].itens[0].valor_total_estimado)

    def test_multi_item_lot_uses_global_ordinal_and_reports_partial_coverage(self) -> None:
        result = AtaSessaoParseResult(
            source_path="ata.pdf",
            generated_at="2026-08-12T10:00:00",
            lotes=[
                LotRecord(1, "ADJUDICADO", "Caneta", itens=[_item("CANETA AZUL")]),
                LotRecord(
                    2,
                    "FRACASSADO",
                    "Materiais",
                    itens=[
                        _item("LAPIS PRETO HB", number="1"),
                        _item("BORRACHA BRANCA GRANDE", number="2"),
                    ],
                ),
            ],
        )
        metadata = reconcile_estimated_values(
            result,
            _sd(
                _sd_item(1, "CANETA AZUL"),
                _sd_item(2, "LAPIS PRETO HB", unitario="2.50"),
                _sd_item(9, "REGUA PLASTICA", unitario="4.00"),
            ),
        )

        failed_items = result.lotes[1].itens
        self.assertEqual(failed_items[0].valor_unitario_estimado, 2.5)
        self.assertEqual(failed_items[0].valor_estimado_correspondencia, "ITEM_GLOBAL")
        self.assertIsNone(failed_items[1].valor_unitario_estimado)
        self.assertEqual(metadata["partially_matched_lots"], 1)
        self.assertEqual(metadata["unmatched_lots"], [2])
        self.assertEqual(metadata["matched_items"], 1)
        self.assertEqual(metadata["unmatched_items"], 1)

    def test_description_fallback_requires_margin_and_quantity_or_unit(self) -> None:
        result = AtaSessaoParseResult(
            source_path="ata.pdf",
            generated_at="2026-08-12T10:00:00",
            lotes=[
                LotRecord(
                    30,
                    "DESERTO",
                    "Equipamento",
                    itens=[_item("MONITOR LED 24 POLEGADAS FULL HD", quantity=3)],
                )
            ],
        )
        metadata = reconcile_estimated_values(
            result,
            _sd(
                _sd_item(
                    4,
                    "MONITOR LED 24 POLEGADAS FULL HD",
                    quantidade="3",
                    unitario="900.00",
                ),
                _sd_item(5, "CABO HDMI 2 METROS", quantidade="3"),
            ),
        )
        item = result.lotes[0].itens[0]
        self.assertEqual(metadata["matched_items"], 1)
        self.assertEqual(item.valor_estimado_correspondencia, "DESCRICAO")
        self.assertEqual(item.valor_estimado_confianca, "MÉDIA")

    def test_ordinal_rejects_strongly_divergent_description_even_when_quantity_and_unit_match(self) -> None:
        result = AtaSessaoParseResult(
            source_path="ata.pdf",
            generated_at="2026-08-12T10:00:00",
            lotes=[
                LotRecord(
                    1,
                    "FRACASSADO",
                    "Cadeira escolar",
                    itens=[_item("CADEIRA ESCOLAR", quantity=1, unit="UN")],
                )
            ],
        )

        metadata = reconcile_estimated_values(
            result,
            _sd(
                _sd_item(
                    1,
                    "TUBO DE CONCRETO ARMADO",
                    quantidade="1",
                    unidade="UN",
                    unitario="900.00",
                )
            ),
        )

        item = result.lotes[0].itens[0]
        self.assertIsNone(item.valor_unitario_estimado)
        self.assertEqual(item.valor_estimado_conciliacao, "AMBIGUO")
        self.assertEqual(metadata["ambiguous_items"], 1)
        self.assertEqual(metadata["ambiguous_lots"], [1])

    def test_sd_overrides_internal_value_only_on_failed_status(self) -> None:
        failed_item = _item("CADEIRA ESCOLAR")
        failed_item.valor_unitario_estimado = 999.0
        failed_item.valor_total_estimado = 999.0
        failed_item.valor_estimado_fonte = "Dossiê"
        successful_item = _item("MESA ESCOLAR")
        successful_item.valor_unitario_estimado = 777.0
        successful_item.valor_total_estimado = 777.0
        successful_item.valor_estimado_fonte = "Dossiê"
        result = AtaSessaoParseResult(
            source_path="ata.pdf",
            generated_at="2026-08-12T10:00:00",
            lotes=[
                LotRecord(1, "FRACASSADO", "Cadeira", itens=[failed_item]),
                LotRecord(2, "ADJUDICADO", "Mesa", itens=[successful_item]),
            ],
        )
        metadata = reconcile_estimated_values(
            result,
            _sd(_sd_item(1, "CADEIRA ESCOLAR", unitario="100.00")),
        )

        self.assertEqual(failed_item.valor_unitario_estimado, 100.0)
        self.assertEqual(
            failed_item.valor_estimado_fonte,
            "Solicitação de Despesa — SD 152/2026",
        )
        self.assertEqual(successful_item.valor_unitario_estimado, 777.0)
        self.assertTrue(any("SD prevaleceu" in warning for warning in metadata["warnings"]))

    def test_zero_failed_lots_still_emits_complete_contract(self) -> None:
        result = AtaSessaoParseResult(
            source_path="ata.pdf",
            generated_at="2026-08-12T10:00:00",
            lotes=[LotRecord(1, "ADJUDICADO", "Mesa", itens=[_item("MESA")])],
        )
        metadata = reconcile_estimated_values(result, _sd(_sd_item(1, "MESA")))

        self.assertEqual(
            set(metadata),
            {
                "source",
                "sd_number",
                "total_failed_lots",
                "fully_matched_lots",
                "partially_matched_lots",
                "unmatched_lots",
                "ambiguous_lots",
                "total_failed_items",
                "matched_items",
                "ambiguous_items",
                "unmatched_items",
                "warnings",
            },
        )
        self.assertEqual(metadata["total_failed_lots"], 0)
        self.assertEqual(metadata["fully_matched_lots"], 0)
        self.assertEqual(metadata["unmatched_lots"], [])
        self.assertEqual(result.to_dict()["estimated_value_reconciliation"], metadata)

    def test_partial_coverage_is_visible_and_has_no_misleading_lot_total(self) -> None:
        result = AtaSessaoParseResult(
            source_path="ata.pdf",
            generated_at="2026-08-12T10:00:00",
            lotes=[
                LotRecord(
                    1,
                    "FRACASSADO",
                    "Materiais escolares",
                    itens=[
                        _item("CADEIRA ESCOLAR", number="1"),
                        _item("MESA ESCOLAR", number="2"),
                    ],
                    motivo_falha="Sem proposta válida.",
                )
            ],
        )
        reconcile_estimated_values(
            result,
            _sd(
                _sd_item(1, "CADEIRA ESCOLAR", unitario="100.00"),
                _sd_item(9, "ARMARIO DE ACO", unitario="800.00"),
            ),
        )
        normalized = normalize_report_data(result)
        self.assertEqual(normalized.malsucedidos[0].valor_estimado_cobertura, "Parcial (1/2)")
        self.assertIsNone(normalized.malsucedidos[0].valor_total_estimado)

        with tempfile.TemporaryDirectory() as tmp_dir:
            xlsx_path = Path(write_reports_workbooks(result, tmp_dir)["malsucedidos_xlsx"])
            pdf_path = Path(
                write_report_pdfs(normalized, tmp_dir)["malsucedidos_pdf"]
            )

            workbook = load_workbook(xlsx_path, data_only=True)
            summary = workbook["Malsucedidos"]
            headers = {cell.value: cell.column for cell in summary[1]}
            self.assertEqual(
                summary.cell(2, headers["Cobertura dos Valores Estimados"]).value,
                "Parcial (1/2)",
            )
            self.assertIsNone(
                summary.cell(2, headers["Valor Total Estimado do Lote"]).value
            )
            item_sheet = workbook["Itens"]
            item_headers = {cell.value: cell.column for cell in item_sheet[1]}
            self.assertEqual(
                item_sheet.cell(2, item_headers["Valor Unitário Estimado"]).value,
                100,
            )
            self.assertEqual(
                item_sheet.cell(2, item_headers["Status da Conciliação"]).value,
                "CONCILIADO",
            )
            self.assertEqual(
                item_sheet.cell(3, item_headers["Status da Conciliação"]).value,
                "NAO_ENCONTRADO",
            )
            reconciliation_sheet = workbook["Conciliação"]
            reconciliation_rows = {
                row[1].value: row[2].value
                for row in reconciliation_sheet.iter_rows(min_row=2)
            }
            self.assertEqual(reconciliation_rows["Número da SD"], "152/2026")
            self.assertEqual(reconciliation_rows["Lotes parcialmente conciliados"], 1)
            self.assertEqual(reconciliation_rows["Parcialmente conciliados"], "1")
            self.assertEqual(reconciliation_rows["Não encontrados"], "1")
            self.assertTrue(
                any(
                    row[0].value == "Avisos" and "nenhum item seguro" in str(row[2].value)
                    for row in reconciliation_sheet.iter_rows(min_row=2)
                )
            )

            with pdfplumber.open(pdf_path) as pdf:
                pdf_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            self.assertIn("Parcial (1/2)", pdf_text)
            self.assertIn("152/2026", pdf_text)
            self.assertIn("Solicitação de", pdf_text)
            self.assertIn("Despesa", pdf_text)
            self.assertNotIn("AVISOS DA CONCILIAÇÃO", pdf_text)
            self.assertNotIn("Sem correspondência", pdf_text)
            self.assertNotIn("Conf.", pdf_text)
            self.assertNotIn("Concil.", pdf_text)

    def test_audit_artifacts_list_ambiguous_and_unmatched_lots(self) -> None:
        result = AtaSessaoParseResult(
            source_path="ata.pdf",
            generated_at="2026-08-12T10:00:00",
            lotes=[
                LotRecord(
                    7,
                    "CANCELADO",
                    "Tubo",
                    itens=[_item("TUBO PVC 100 MM", catmat="123456")],
                    motivo_falha="Cancelado pela autoridade competente.",
                ),
                LotRecord(
                    8,
                    "DESERTO",
                    "Projetor",
                    itens=[_item("PROJETOR MULTIMIDIA PROFISSIONAL")],
                    motivo_falha="Lote sem propostas.",
                ),
            ],
        )
        reconciliation = reconcile_estimated_values(
            result,
            _sd(
                _sd_item(70, "TUBO PVC 100 MM", catmat="123456"),
                _sd_item(71, "TUBO PVC 100 MM REFORCADO", catmat="123456"),
                _sd_item(99, "ARMARIO DE ACO"),
            ),
        )

        with tempfile.TemporaryDirectory() as tmp_dir:
            xlsx_path = Path(write_reports_workbooks(result, tmp_dir)["malsucedidos_xlsx"])
            pdf_path = Path(
                write_report_pdfs(normalize_report_data(result), tmp_dir)[
                    "malsucedidos_pdf"
                ]
            )
            institutional_pdf_path = Path(
                write_ata_institucional_pdf(
                    result,
                    normalize_report_data(result),
                    tmp_dir,
                )["ata_institucional_pdf"]
            )
            workbook = load_workbook(xlsx_path, data_only=True)
            audit_sheet = workbook["Conciliação"]
            audit_values = {
                row[1].value: row[2].value for row in audit_sheet.iter_rows(min_row=2)
            }
            self.assertEqual(audit_values["Ambíguos"], "7")
            self.assertEqual(audit_values["Não encontrados"], "8")
            self.assertGreaterEqual(audit_values["Itens ambíguos"], 1)
            self.assertGreaterEqual(audit_values["Itens não encontrados"], 1)

            with pdfplumber.open(pdf_path) as pdf:
                pdf_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            self.assertNotIn("AVISOS DA CONCILIAÇÃO", pdf_text)
            self.assertNotIn("Correspondência ambígua", pdf_text)
            self.assertNotIn("Sem correspondência", pdf_text)

            with pdfplumber.open(institutional_pdf_path) as pdf:
                institutional_pdf_text = "\n".join(
                    page.extract_text() or "" for page in pdf.pages
                )
            for warning in reconciliation["warnings"]:
                self.assertNotIn(warning, institutional_pdf_text)

    def test_complete_reconciliation_omits_pdf_warning_section(self) -> None:
        result = AtaSessaoParseResult(
            source_path="ata.pdf",
            generated_at="2026-08-12T10:00:00",
            lotes=[
                LotRecord(
                    1,
                    "FRACASSADO",
                    "Cadeira escolar",
                    itens=[_item("CADEIRA ESCOLAR")],
                    motivo_falha="Sem proposta válida.",
                )
            ],
        )
        metadata = reconcile_estimated_values(
            result,
            _sd(_sd_item(1, "CADEIRA ESCOLAR", unitario="100.00")),
        )
        self.assertEqual(metadata["warnings"], [])

        with tempfile.TemporaryDirectory() as tmp_dir:
            artifacts = write_reports_workbooks(result, tmp_dir)
            workbook = load_workbook(artifacts["malsucedidos_xlsx"], data_only=True)
            audit_sheet = workbook["Conciliação"]
            audit_values = {
                row[1].value: row[2].value for row in audit_sheet.iter_rows(min_row=2)
            }
            self.assertEqual(audit_values["Ambíguos"], "Nenhum")
            self.assertEqual(audit_values["Não encontrados"], "Nenhum")

            pdf_path = Path(
                write_report_pdfs(normalize_report_data(result), tmp_dir)[
                    "malsucedidos_pdf"
                ]
            )
            with pdfplumber.open(pdf_path) as pdf:
                pdf_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            self.assertNotIn("AVISOS DA CONCILIAÇÃO", pdf_text)
            self.assertNotIn("Confiança do Valor Estimado", pdf_text)
            self.assertNotIn("Status da Conciliação", pdf_text)
            self.assertNotIn("Correspondência", pdf_text)


    def test_item_detail_heading_is_not_orphaned_at_page_footer(self) -> None:
        result = AtaSessaoParseResult(
            source_path="ata.pdf",
            generated_at="2026-08-12T10:00:00",
            lotes=[
                LotRecord(
                    1,
                    "FRACASSADO",
                    "Tela de protecao",
                    itens=[
                        _item(
                            "TELA DE PROTECAO EM FIBRA DE VIDRO PARA JANELAS",
                            catmat="9900515461",
                            quantity=30,
                            unit="ROL",
                        )
                    ],
                    motivo_falha=(
                        "A proposta foi desclassificada por descumprimento material "
                        "das exigencias do instrumento convocatorio. "
                    )
                    * 22,
                )
            ],
        )
        reconcile_estimated_values(
            result,
            _sd(
                _sd_item(
                    1,
                    "TELA DE PROTECAO EM FIBRA DE VIDRO PARA JANELAS",
                    catmat="9900515461",
                    quantidade="30",
                    unidade="ROL",
                    unitario="267.32",
                )
            ),
        )

        with tempfile.TemporaryDirectory() as tmp_dir:
            pdf_path = Path(
                write_report_pdfs(normalize_report_data(result), tmp_dir)[
                    "malsucedidos_pdf"
                ]
            )
            with pdfplumber.open(pdf_path) as pdf:
                pages = [page.extract_text() or "" for page in pdf.pages]

        heading_pages = {
            index for index, text in enumerate(pages) if "DETALHAMENTO DO ITEM" in text
        }
        field_pages = {
            index for index, text in enumerate(pages) if "CATMAT/CATSER" in text
        }
        self.assertTrue(heading_pages)
        self.assertTrue(heading_pages & field_pages)
        self.assertIn(len(pages) - 1, field_pages)


if __name__ == "__main__":
    unittest.main()
