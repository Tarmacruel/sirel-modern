from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import AtaSessaoParseResult, LotItemData, LotRecord

CURRENCY_FORMAT = '[$R$-416] #,##0.00'
HEADER_FILL = PatternFill(fill_type='solid', fgColor='2440A7')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
BODY_ALIGNMENT = Alignment(vertical='top', wrap_text=True, horizontal='left')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def _first_item(lot: LotRecord) -> LotItemData | None:
    return lot.itens[0] if lot.itens else None


def _sum_quantidade(lot: LotRecord) -> float | None:
    total = sum((item.quantidade or 0) for item in lot.itens)
    return total if total > 0 else None


def _sum_valor_total(lot: LotRecord) -> float | None:
    total = sum((item.valor_total or 0) for item in lot.itens)
    return total if total > 0 else None


def _participant_rows(lots: Iterable[LotRecord]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for lot in lots:
        for p in lot.participantes:
            rows.append(
                {
                    'Nº Lote': lot.numero_lote,
                    'Status do Lote': lot.status,
                    'Seção': p.section,
                    'Colocação': p.ranking,
                    'Razão Social': p.razao_social,
                    'Nº Participante': p.participante_numero,
                    'CNPJ/CPF': p.documento,
                    'Oferta Inicial': p.oferta_inicial,
                    'Oferta Final': p.oferta_final,
                    'Dif. (%)': None if p.diferenca_percentual is None else f'{p.diferenca_percentual:.2f}%',
                    'ME/EPP': 'Sim' if p.me_epp else 'Não' if p.me_epp is False else '',
                }
            )
    return rows


def _operacionais_rows(lots: Iterable[LotRecord]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for lot in lots:
        item = _first_item(lot)
        rows.append(
            {
                'Nº Lote': lot.numero_lote,
                'Descrição do Lote/1º Item': (item.descricao if item and item.descricao else None) or lot.titulo,
                'Total de Itens': len(lot.itens),
                'Quantidade Total': _sum_quantidade(lot),
                'Valor Total do Lote': _sum_valor_total(lot),
                'Marca (1º item)': item.marca if item and item.marca else '',
                'Modelo (1º item)': item.modelo if item and item.modelo else '',
                'Status': lot.status,
                'Vencedor': lot.vencedor or '',
                'CNPJ Vencedor': lot.cnpj_vencedor or '',
                'Melhor Lance (R$)': lot.melhor_lance,
            }
        )
    return rows


def _malsucedidos_rows(lots: Iterable[LotRecord]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for lot in lots:
        item = _first_item(lot)
        matched_items = [
            current
            for current in lot.itens
            if current.valor_unitario_estimado is not None
            and current.valor_total_estimado is not None
        ]
        coverage_complete = bool(lot.itens) and len(matched_items) == len(lot.itens)
        estimated_lot_total = (
            sum(float(current.valor_total_estimado or 0) for current in lot.itens)
            if coverage_complete
            else None
        )
        provenance_item = matched_items[0] if matched_items else item
        marca_modelo = ' / '.join(
            part
            for part in [
                item.marca if item and item.marca else None,
                item.modelo if item and item.modelo else None,
            ]
            if part
        )
        rows.append(
            {
                'Nº Lote': lot.numero_lote,
                'Descrição do Lote/1º Item': (item.descricao if item and item.descricao else None) or lot.titulo,
                'Total de Itens': len(lot.itens),
                'Valor Unitário Estimado (lote unitário)': (
                    item.valor_unitario_estimado if item and len(lot.itens) == 1 else None
                ),
                'Valor Total Estimado do Lote': estimated_lot_total,
                'Cobertura dos Valores Estimados': (
                    f"Completa ({len(matched_items)}/{len(lot.itens)})"
                    if coverage_complete
                    else f"Parcial ({len(matched_items)}/{len(lot.itens)})"
                    if matched_items
                    else f"Não conciliada (0/{len(lot.itens)})"
                ),
                'Itens Conciliados': f"{len(matched_items)}/{len(lot.itens)}",
                'Processo Fonte do Valor Estimado': provenance_item.valor_estimado_processo_fonte if provenance_item else '',
                'Fonte do Valor Estimado': provenance_item.valor_estimado_fonte if provenance_item else '',
                'Confiança do Valor Estimado': provenance_item.valor_estimado_confianca if provenance_item else '',
                'Marca/Modelo (1º item)': marca_modelo,
                'Status': lot.status,
                'Motivo da Falha': lot.motivo_falha or '',
            }
        )
    return rows


def _item_rows(lots: Iterable[LotRecord]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for lot in lots:
        if not lot.itens:
            rows.append(
                {
                    'Nº Lote': lot.numero_lote,
                    'Status do Lote': lot.status,
                    'Item': '',
                    'CATMAT/CATSER': '',
                    'Descrição': lot.titulo,
                    'Unidade': '',
                    'Quantidade': None,
                    'Valor Unitário': None,
                    'Valor Total': None,
                    'Valor Unitário Estimado': None,
                    'Valor Total Estimado': None,
                    'Processo Fonte do Valor Estimado': '',
                    'Fonte do Valor Estimado': '',
                    'Confiança do Valor Estimado': '',
                    'Status da Conciliação': '',
                    'Correspondência': '',
                    'Marca': '',
                    'Modelo': '',
                    'Vencedor': lot.vencedor or '',
                    'CNPJ Vencedor': lot.cnpj_vencedor or '',
                }
            )
            continue

        for item in lot.itens:
            rows.append(
                {
                    'Nº Lote': lot.numero_lote,
                    'Status do Lote': lot.status,
                    'Item': item.item_numero or '',
                    'CATMAT/CATSER': item.catmat_catser or '',
                    'Descrição': item.descricao or lot.titulo,
                    'Unidade': item.unidade or '',
                    'Quantidade': item.quantidade,
                    'Valor Unitário': item.valor_unitario,
                    'Valor Total': item.valor_total,
                    'Valor Unitário Estimado': item.valor_unitario_estimado,
                    'Valor Total Estimado': item.valor_total_estimado,
                    'Processo Fonte do Valor Estimado': item.valor_estimado_processo_fonte or '',
                    'Fonte do Valor Estimado': item.valor_estimado_fonte or '',
                    'Confiança do Valor Estimado': item.valor_estimado_confianca or '',
                    'Status da Conciliação': item.valor_estimado_conciliacao or '',
                    'Correspondência': item.valor_estimado_correspondencia or '',
                    'Marca': item.marca or '',
                    'Modelo': item.modelo or '',
                    'Vencedor': lot.vencedor or '',
                    'CNPJ Vencedor': lot.cnpj_vencedor or '',
                }
            )
    return rows


def _partially_matched_lot_numbers(lots: Iterable[LotRecord]) -> list[int]:
    numbers: list[int] = []
    for lot in lots:
        total_items = len(lot.itens)
        matched_items = sum(
            item.valor_unitario_estimado is not None
            and item.valor_total_estimado is not None
            for item in lot.itens
        )
        if total_items > 0 and 0 < matched_items < total_items:
            numbers.append(lot.numero_lote)
    return sorted(set(numbers))


def _lot_numbers_text(values: object) -> str:
    if not isinstance(values, (list, tuple, set)):
        return 'Nenhum'
    normalized: list[int] = []
    for value in values:
        try:
            number = int(value)
        except (TypeError, ValueError):
            continue
        if number > 0:
            normalized.append(number)
    return ', '.join(str(number) for number in sorted(set(normalized))) or 'Nenhum'


def _reconciliation_rows(result: AtaSessaoParseResult) -> list[dict[str, object]]:
    reconciliation = result.estimated_value_reconciliation
    if not reconciliation:
        return [
            {
                'Seção': 'Disponibilidade',
                'Indicador': 'Status',
                'Valor': 'Conciliação com Solicitação de Despesa não executada.',
            }
        ]

    metric_labels: list[tuple[str, str]] = [
        ('source', 'Fonte'),
        ('sd_number', 'Número da SD'),
        ('total_failed_lots', 'Total de lotes malsucedidos'),
        ('fully_matched_lots', 'Lotes totalmente conciliados'),
        ('partially_matched_lots', 'Lotes parcialmente conciliados'),
        ('total_failed_items', 'Total de itens malsucedidos'),
        ('matched_items', 'Itens conciliados'),
        ('ambiguous_items', 'Itens ambíguos'),
        ('unmatched_items', 'Itens não encontrados'),
    ]
    rows: list[dict[str, object]] = [
        {
            'Seção': 'Métricas',
            'Indicador': label,
            'Valor': reconciliation.get(key),
        }
        for key, label in metric_labels
    ]
    rows.extend(
        [
            {
                'Seção': 'Lotes afetados',
                'Indicador': 'Parcialmente conciliados',
                'Valor': _lot_numbers_text(_partially_matched_lot_numbers(result.malsucedidos)),
            },
            {
                'Seção': 'Lotes afetados',
                'Indicador': 'Ambíguos',
                'Valor': _lot_numbers_text(reconciliation.get('ambiguous_lots')),
            },
            {
                'Seção': 'Lotes afetados',
                'Indicador': 'Não encontrados',
                'Valor': _lot_numbers_text(reconciliation.get('unmatched_lots')),
            },
        ]
    )
    warnings = reconciliation.get('warnings')
    if isinstance(warnings, list):
        rows.extend(
            {
                'Seção': 'Avisos',
                'Indicador': f'Aviso {index}',
                'Valor': str(warning),
            }
            for index, warning in enumerate(warnings, start=1)
            if str(warning).strip()
        )
    return rows


def _format_worksheet(worksheet, currency_columns: set[str]) -> None:
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = BODY_ALIGNMENT
            cell.border = THIN_BORDER
            header_val = worksheet.cell(row=1, column=cell.column).value
            if header_val in currency_columns and isinstance(cell.value, (int, float)):
                cell.number_format = CURRENCY_FORMAT

    for col_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col_cells[0].column)

        for cell in col_cells:
            try:
                length = len(str(cell.value or ''))
                if length > max_length:
                    max_length = length
            except Exception:
                continue

        adjusted_width = min(max(max_length + 2, 14), 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _write_sheet(
    writer: pd.ExcelWriter,
    *,
    sheet_name: str,
    rows: list[dict[str, object]],
    currency_columns: set[str],
    empty_message: str,
) -> None:
    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame([{'Aviso': empty_message}])

    df.to_excel(writer, sheet_name=sheet_name, index=False)
    _format_worksheet(writer.sheets[sheet_name], currency_columns)


def write_reports_workbooks(result: AtaSessaoParseResult, output_dir: str | Path) -> dict[str, str]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    andamento_path = output_dir / 'Relatorio_EmAndamento.xlsx'
    adj_path = output_dir / 'Relatorio_Adjudicados.xlsx'
    recursal_path = output_dir / 'Relatorio_FaseRecursal.xlsx'
    mal_path = output_dir / 'Relatorio_MalSucedidos.xlsx'

    with pd.ExcelWriter(andamento_path, engine='openpyxl') as writer:
        _write_sheet(
            writer,
            sheet_name='Em andamento',
            rows=_operacionais_rows(result.em_andamento),
            currency_columns={'Valor Total do Lote', 'Melhor Lance (R$)'},
            empty_message='Nenhum registro encontrado para este relatório.',
        )
        _write_sheet(
            writer,
            sheet_name='Itens',
            rows=_item_rows(result.em_andamento),
            currency_columns={'Valor Unitário', 'Valor Total', 'Valor Unitário Estimado'},
            empty_message='Nenhum item registrado.',
        )
        _write_sheet(
            writer,
            sheet_name='Participantes',
            rows=_participant_rows(result.em_andamento),
            currency_columns={'Oferta Inicial', 'Oferta Final'},
            empty_message='Nenhum participante registrado.',
        )

    with pd.ExcelWriter(adj_path, engine='openpyxl') as writer:
        _write_sheet(
            writer,
            sheet_name='Adjudicados',
            rows=_operacionais_rows(result.adjudicados),
            currency_columns={'Valor Total do Lote', 'Melhor Lance (R$)'},
            empty_message='Nenhum registro encontrado para este relatório.',
        )
        _write_sheet(
            writer,
            sheet_name='Itens',
            rows=_item_rows(result.adjudicados),
            currency_columns={'Valor Unitário', 'Valor Total', 'Valor Unitário Estimado'},
            empty_message='Nenhum item registrado.',
        )
        _write_sheet(
            writer,
            sheet_name='Participantes',
            rows=_participant_rows(result.adjudicados),
            currency_columns={'Oferta Inicial', 'Oferta Final'},
            empty_message='Nenhum participante registrado.',
        )

    with pd.ExcelWriter(recursal_path, engine='openpyxl') as writer:
        _write_sheet(
            writer,
            sheet_name='Fase recursal',
            rows=_operacionais_rows(result.fase_recursal),
            currency_columns={'Valor Total do Lote', 'Melhor Lance (R$)'},
            empty_message='Nenhum registro encontrado para este relatório.',
        )
        _write_sheet(
            writer,
            sheet_name='Itens',
            rows=_item_rows(result.fase_recursal),
            currency_columns={'Valor Unitário', 'Valor Total', 'Valor Unitário Estimado'},
            empty_message='Nenhum item registrado.',
        )
        _write_sheet(
            writer,
            sheet_name='Participantes',
            rows=_participant_rows(result.fase_recursal),
            currency_columns={'Oferta Inicial', 'Oferta Final'},
            empty_message='Nenhum participante registrado.',
        )

    with pd.ExcelWriter(mal_path, engine='openpyxl') as writer:
        _write_sheet(
            writer,
            sheet_name='Malsucedidos',
            rows=_malsucedidos_rows(result.malsucedidos),
            currency_columns={'Valor Unitário Estimado (lote unitário)', 'Valor Total Estimado do Lote'},
            empty_message='Nenhum registro encontrado para este relatório.',
        )
        _write_sheet(
            writer,
            sheet_name='Itens',
            rows=_item_rows(result.malsucedidos),
            currency_columns={'Valor Unitário', 'Valor Total', 'Valor Unitário Estimado', 'Valor Total Estimado'},
            empty_message='Nenhum item registrado.',
        )
        _write_sheet(
            writer,
            sheet_name='Participantes',
            rows=_participant_rows(result.malsucedidos),
            currency_columns={'Oferta Inicial', 'Oferta Final'},
            empty_message='Nenhum participante registrado.',
        )
        _write_sheet(
            writer,
            sheet_name='Conciliação',
            rows=_reconciliation_rows(result),
            currency_columns=set(),
            empty_message='Nenhuma informação de conciliação registrada.',
        )

    return {
        'em_andamento_xlsx': str(andamento_path),
        'adjudicados_xlsx': str(adj_path),
        'fase_recursal_xlsx': str(recursal_path),
        'malsucedidos_xlsx': str(mal_path),
    }
